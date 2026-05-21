"""
REQ-006 doc_parser — adapters/parsers/xlsx_parser.py

XlsxParser
openpyxl 기반 Excel 파서

처리 항목:
    - 시트별 테이블 블록 변환
    - SheetMeta 생성 (시트명, 행/열 수)
    - styles.xml count 속성 호환성 문제 자동 복구
    - 3층 구조 처리:
        1층: 셀 데이터 (flat rows, 청킹용)
        2층: 병합셀 raw_grid + normalized_headers
        3층: 차트/이미지 시각 객체 감지 (vision 트리거용)

주의:
    read_only=True 모드에서는 merged_cells / _charts / _images 접근 불가.
    병합셀 감지를 위해 data_only=True (read_only=False) 모드로 로딩.
    대용량 파일의 경우 60초 성능 기준 내 처리 가능한 수준.
"""
from __future__ import annotations

import io
import json
import re
import zipfile
from uuid import uuid4

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from common_schemas.document import (
    ContentBlock,
    DocumentBlock,
    FileMeta,
    ParserMeta,
    SheetMeta,
    SourceRef,
)

from doc_parser.domain.ports.parser_port import ParserPort

# 헤더 행으로 판단할 최대 행 수
# 이 범위 내에서 병합셀이 있으면 다층 헤더로 간주
_MAX_HEADER_ROWS = 3


class XlsxParser(ParserPort):
    """Excel 파서 구현체.

    지원 MIME 타입:
        application/vnd.openxmlformats-officedocument.spreadsheetml.sheet

    Raises:
        RuntimeError: 파일 손상 또는 읽기 실패 E0202
        RuntimeError: 텍스트 추출 실패 E0203
    """

    MIME_TYPE = (
        "application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet"
    )

    # ──────────────────────────────────────────
    # ParserPort 구현
    # ──────────────────────────────────────────

    def parse(
        self,
        file_path: str,
        file_meta: FileMeta,
    ) -> DocumentBlock:
        """XLSX 파싱 → DocumentBlock 반환.

        병합셀이 없는 시트: 1층 flat rows → table 필드에 list[list[str]]
        병합셀이 있는 시트: table=data_rows(flat), metadata=병합셀 구조 dict

        Args:
            file_path: XLSX 파일 경로
            file_meta: 파일 메타데이터

        Returns:
            DocumentBlock: 파싱된 문서 블록 (시트별 table 블록)

        Raises:
            RuntimeError: 파일 손상 (E0202), 텍스트 추출 실패 (E0203)
        """
        try:
            wb = self._load_workbook_safe(file_path)
        except Exception as e:
            raise RuntimeError(f"E0202: XLSX 파일 읽기 실패 — {e}") from e

        try:
            blocks: list[ContentBlock] = []
            sheet_metas: list[SheetMeta] = []
            block_index = 0

            for sheet_index, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                has_merged = bool(list(ws.merged_cells.ranges))

                if has_merged:
                    # ── 2층+3층: 병합셀 시트 ──
                    block = self._build_layered_block(
                        ws=ws,
                        sheet_index=sheet_index,
                        sheet_name=sheet_name,
                        block_index=block_index,
                    )
                    if block is None:
                        continue

                    # SheetMeta: metadata의 raw_grid 기준 집계
                    raw_grid = block.metadata["raw_grid"]  # type: ignore[index]
                    row_count = len(raw_grid)
                    col_count = max(len(r) for r in raw_grid) if raw_grid else 0
                else:
                    # ── 1층: 기존 flat rows ──
                    rows = self._extract_rows(ws)
                    if not rows:
                        continue

                    row_count = len(rows)
                    col_count = max(len(r) for r in rows) if rows else 0

                    block = ContentBlock(
                        block_id=uuid4(),
                        block_type="table",
                        content=None,
                        page=sheet_index + 1,
                        table=rows,
                        source_ref=SourceRef(
                            page=sheet_index + 1,
                            sheet_name=sheet_name,
                            block_index=block_index,
                        ),
                    )

                sheet_metas.append(
                    SheetMeta(
                        sheet_name=sheet_name,
                        row_count=row_count,
                        col_count=col_count,
                    )
                )
                blocks.append(block)
                block_index += 1

            wb.close()

            updated_file_meta = file_meta.model_copy(
                update={"sheet_meta": sheet_metas}
            )

            return DocumentBlock(
                document_id=uuid4(),
                file_meta=updated_file_meta,
                parser=ParserMeta(
                    parser_name="XlsxParser",
                    parser_version="1.0.0",
                ),
                blocks=blocks,
            )

        except Exception as e:
            raise RuntimeError(f"E0203: XLSX 텍스트 추출 실패 — {e}") from e

    def supports(self, mime_type: str) -> bool:
        return mime_type == self.MIME_TYPE

    # ──────────────────────────────────────────
    # Private — 3층 구조 빌더
    # ──────────────────────────────────────────

    def _build_layered_block(
        self,
        ws: Worksheet,
        sheet_index: int,
        sheet_name: str,
        block_index: int,
    ) -> ContentBlock | None:
        """병합셀 시트 → 2층+3층 ContentBlock 생성.

        table 필드 구조:
            [
                {
                    "raw_grid": list[list[str]],          # 원본 셀 그대로
                    "normalized_headers": list[str],       # 부모_자식 flatten
                    "merged_cells": list[dict],            # 병합 범위 + 값
                    "data_rows": list[list[str]],          # 헤더 이후 데이터
                    "has_charts": bool,                    # 3층: 차트 존재
                    "has_images": bool,                    # 3층: 이미지 존재
                }
            ]

        Returns:
            ContentBlock | None: 데이터 없으면 None
        """
        raw_grid = self._extract_rows(ws)
        if not raw_grid:
            return None

        merged_cells_info, merge_map = self._extract_merged_grid(ws)
        header_row_count = self._detect_header_rows(ws, merge_map)
        normalized_headers = self._normalize_headers(ws, merge_map, header_row_count)
        data_rows = raw_grid[header_row_count:]
        has_charts, has_images = self._detect_visual_objects(ws)

        layered: dict = {
            "raw_grid": raw_grid,
            "normalized_headers": normalized_headers,
            "merged_cells": merged_cells_info,
            "data_rows": data_rows,
            "has_charts": has_charts,
            "has_images": has_images,
        }

        return ContentBlock(
            block_id=uuid4(),
            block_type="table",
            content=None,
            page=sheet_index + 1,
            table=data_rows if data_rows else raw_grid,  # flat rows — 청킹용
            metadata=layered,                             # 병합셀 구조 — QualityGate/Vision용
            source_ref=SourceRef(
                page=sheet_index + 1,
                sheet_name=sheet_name,
                block_index=block_index,
            ),
        )

    # ──────────────────────────────────────────
    # Private — 2층: 병합셀
    # ──────────────────────────────────────────

    def _extract_merged_grid(
        self,
        ws: Worksheet,
    ) -> tuple[list[dict], dict[tuple[int, int], str]]:
        """병합셀 정보 추출.

        Returns:
            merged_cells_info: [{"range": "B1:C1", "value": "주문"}, ...]
            merge_map: {(row, col): 대표셀_값} — normalized_headers 생성용
        """
        merged_cells_info: list[dict] = []
        merge_map: dict[tuple[int, int], str] = {}

        for mr in ws.merged_cells.ranges:
            val = ws.cell(mr.min_row, mr.min_col).value
            val_str = str(val) if val is not None else ""

            merged_cells_info.append({
                "range": str(mr),
                "value": val_str,
            })

            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merge_map[(r, c)] = val_str

        return merged_cells_info, merge_map

    def _detect_header_rows(
        self,
        ws: Worksheet,
        merge_map: dict[tuple[int, int], str],
    ) -> int:
        """헤더 행 수 감지.

        B1:C1 처럼 1행 내 병합의 경우 max_merged_row=1 이므로
        병합 범위만으로는 다층 헤더를 감지할 수 없음.

        대신 다음 조건으로 순차 판단:
            - 다음 행이 비어있으면 → 현재가 마지막 헤더
            - 다음 행 값의 과반수가 숫자면 → 데이터 행 → 현재가 마지막 헤더
            - 현재 행에 병합 gap(None)이 있으면 → 다음 행도 헤더 가능성

        Returns:
            int: 헤더 행 수 (1 이상, 최대 _MAX_HEADER_ROWS)
        """
        if not merge_map:
            return 1

        max_col = ws.max_column or 1

        for header_count in range(1, _MAX_HEADER_ROWS):
            next_row_num = header_count + 1
            next_row_vals = [
                ws.cell(next_row_num, c).value
                for c in range(1, max_col + 1)
            ]

            # 다음 행이 비어있으면 현재가 마지막 헤더
            if not any(v is not None for v in next_row_vals):
                return header_count

            # 다음 행 값의 과반수가 숫자 → 데이터 행
            non_none_vals = [v for v in next_row_vals if v is not None]
            numeric_count = sum(
                1 for v in non_none_vals if isinstance(v, (int, float))
            )
            if non_none_vals and numeric_count / len(non_none_vals) > 0.5:
                return header_count

            # 현재 행에 병합 gap(None)이 있으면 다음 행도 서브헤더일 수 있음
            current_row_vals = [
                ws.cell(header_count, c).value
                for c in range(1, max_col + 1)
            ]
            has_merge_gap = any(
                v is None and (header_count, c + 1) in merge_map
                for c, v in enumerate(current_row_vals)
            )
            if has_merge_gap:
                continue

            return header_count

        return _MAX_HEADER_ROWS

    def _normalize_headers(
        self,
        ws: Worksheet,
        merge_map: dict[tuple[int, int], str],
        header_row_count: int,
    ) -> list[str]:
        """병합셀 기반 normalized_headers 생성.

        헤더 1행:  셀 값 그대로
        헤더 2행:  부모헤더_자식헤더 형태로 flatten
            예) B1:C1 병합 "주문", B2="주문수", C2="매출"
                → "주문_주문수", "주문_매출"

        주의:
            자식 헤더(서브헤더 행)는 merge_map을 사용하지 않고 직접 읽음.
            merge_map은 1행 병합 대표값만 담고 있어서, 자식 행 값은 덮어쓰면 안 됨.

        Args:
            ws: 워크시트
            merge_map: {(row, col): 대표셀_값} — 1행 병합 위치 파악용
            header_row_count: 헤더 행 수

        Returns:
            list[str]: 컬럼별 normalized 헤더
        """
        max_col = ws.max_column or 0
        if max_col == 0:
            return []

        normalized: list[str] = []

        for col in range(1, max_col + 1):
            if header_row_count == 1:
                # 병합 대표값 우선, 없으면 직접 읽기
                val = merge_map.get((1, col))
                if val is None:
                    raw = ws.cell(1, col).value
                    val = str(raw) if raw is not None else ""
                normalized.append(val)
            else:
                # 1행 parent: 병합 대표값 우선, 없으면 직접 읽기
                parent = merge_map.get((1, col))
                if parent is None:
                    raw = ws.cell(1, col).value
                    parent = str(raw) if raw is not None else ""

                # 마지막 헤더 행 child: 항상 직접 읽기 (merge_map 사용 안 함)
                raw_child = ws.cell(header_row_count, col).value
                child = str(raw_child) if raw_child is not None else ""

                if child and child != parent:
                    normalized.append(f"{parent}_{child}")
                else:
                    normalized.append(parent)

        return normalized

    # ──────────────────────────────────────────
    # Private — 3층: 시각 객체 감지
    # ──────────────────────────────────────────

    def _detect_visual_objects(
        self,
        ws: Worksheet,
    ) -> tuple[bool, bool]:
        """차트 / 이미지 존재 여부 감지.

        openpyxl normal 모드에서만 접근 가능.
        read_only 모드에서는 항상 (False, False) 반환.

        Returns:
            (has_charts, has_images)
        """
        try:
            has_charts = bool(getattr(ws, "_charts", []))
            has_images = bool(getattr(ws, "_images", []))
            return has_charts, has_images
        except Exception:
            return False, False

    # ──────────────────────────────────────────
    # Private — 공통 유틸
    # ──────────────────────────────────────────

    def _load_workbook_safe(
        self,
        file_path: str,
    ) -> openpyxl.Workbook:
        """XLSX 로딩 — styles.xml count 속성 호환성 문제 자동 복구.

        data_only=True, read_only=False 로 로딩.
        (병합셀 / 차트 / 이미지 접근을 위해 read_only 비활성화)

        일부 XLSX 파일의 styles.xml에 openpyxl이 인식하지 못하는
        count 속성이 포함된 경우, 해당 속성을 제거한 뒤 재시도.
        데이터 셀 값에는 영향 없음.

        Args:
            file_path: XLSX 파일 경로

        Returns:
            openpyxl.Workbook

        Raises:
            Exception: 복구 후에도 로딩 실패 시 원본 예외 전파
        """
        try:
            return openpyxl.load_workbook(file_path, data_only=True)
        except Exception:
            # styles.xml 에서 count 속성 제거 후 재시도
            with open(file_path, "rb") as f:
                raw = f.read()

            fixed_buf = io.BytesIO()
            with zipfile.ZipFile(io.BytesIO(raw), "r") as zin:
                with zipfile.ZipFile(
                    fixed_buf, "w", zipfile.ZIP_DEFLATED
                ) as zout:
                    for item in zin.namelist():
                        if item == "xl/styles.xml":
                            xml = zin.read(item).decode("utf-8")
                            xml = re.sub(r'\s+count="\d+"', "", xml)
                            zout.writestr(item, xml)
                        else:
                            zout.writestr(item, zin.read(item))

            fixed_buf.seek(0)
            return openpyxl.load_workbook(fixed_buf, data_only=True)

    def _extract_rows(
        self,
        ws: Worksheet,
    ) -> list[list[str]]:
        """워크시트에서 행 데이터 추출.

        빈 행은 제외.
        셀 값은 문자열로 변환, None → 빈 문자열.
        """
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cells):
                rows.append(cells)
        return rows
